import json, sys
from openpyxl import load_workbook

for path in sys.argv[1:]:
    wb = load_workbook(path, read_only=True, data_only=False)
    print(f"\n=== {path} ===")
    for ws in wb.worksheets:
        print(json.dumps({"sheet": ws.title, "max_row": ws.max_row, "max_column": ws.max_column}, ensure_ascii=False))
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True):
            print(json.dumps(list(row[:25]), ensure_ascii=False, default=str))
