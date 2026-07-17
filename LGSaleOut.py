"""LGSaleOut Excel import server.

Run:
    python LGSaleOut.py

Optional environment variable:
    LGSALEOUT_CONNECTION_STRING

Example:
    DRIVER={ODBC Driver 17 for SQL Server};SERVER=127.0.0.1,49172;DATABASE=LGSaleOut;UID=Tim;PWD=...;
"""

from __future__ import annotations

import os
import re
import secrets
import socket
import tempfile
import threading
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pyodbc
from flask import Flask, jsonify, request, send_from_directory
from werkzeug.utils import secure_filename


BASE_DIR = Path(__file__).resolve().parent
EXPECTED_HEADERS = ["TW CODE", "簡稱", "產品別", "產品別2", "型號", "陳列", "期末", "不計", "業務員"]
MAX_FILE_SIZE = 20 * 1024 * 1024
TOKEN_TTL_SECONDS = 30 * 60

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_FILE_SIZE


@dataclass
class PendingImport:
    path: Path
    original_name: str
    period_end: date
    rows: list[dict[str, Any]]
    warnings: list[dict[str, Any]] = field(default_factory=list)
    created_at: datetime = field(default_factory=datetime.now)


pending_imports: dict[str, PendingImport] = {}
pending_lock = threading.Lock()


def connection_string() -> str:
    configured = os.getenv("LGSALEOUT_CONNECTION_STRING")
    if configured:
        return configured
    return (
        "DRIVER={ODBC Driver 17 for SQL Server};"
        "SERVER=127.0.0.1,49172;"
        "DATABASE=LGSaleOut;"
        "UID=Tim;"
        "PWD=561202;"
        "Encrypt=no;"
        "TrustServerCertificate=yes;"
        "Connection Timeout=5;"
    )


def db_connect() -> pyodbc.Connection:
    return pyodbc.connect(connection_string(), timeout=8)


def local_network_ip() -> str | None:
    """Best-effort LAN address detection without sending application data."""
    probe = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        probe.connect(("8.8.8.8", 80))
        address = probe.getsockname()[0]
        return address if not address.startswith("127.") else None
    except OSError:
        try:
            addresses = socket.gethostbyname_ex(socket.gethostname())[2]
            return next((item for item in addresses if not item.startswith("127.")), None)
        except OSError:
            return None
    finally:
        probe.close()


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_integer(value: Any, field_name: str, row_number: int, nullable: bool) -> int | None:
    if value is None or clean_text(value) is None:
        if nullable:
            return None
        raise ValueError(f"第 {row_number} 列：{field_name}不可空白")
    if isinstance(value, bool):
        raise ValueError(f"第 {row_number} 列：{field_name}必須是整數")
    if isinstance(value, float) and not value.is_integer():
        raise ValueError(f"第 {row_number} 列：{field_name}必須是整數")
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"第 {row_number} 列：{field_name}必須是整數") from exc


def period_from_filename(filename: str) -> date | None:
    match = re.search(r"(?<!\d)(20\d{2})(0[1-9]|1[0-2])(?!\d)", filename)
    if not match:
        return None
    year, month = int(match.group(1)), int(match.group(2))
    if month == 12:
        return date(year, 12, 31)
    return date(year, month + 1, 1) - __import__("datetime").timedelta(days=1)


def parse_period(value: str | None, filename: str) -> date:
    if value:
        try:
            year, month = map(int, value.split("-"))
            if month == 12:
                return date(year, 12, 31)
            return date(year, month + 1, 1) - __import__("datetime").timedelta(days=1)
        except (TypeError, ValueError):
            raise ValueError("月份格式不正確")
    inferred = period_from_filename(filename)
    if inferred is None:
        raise ValueError("無法從檔名辨識月份，請在畫面選擇月份")
    return inferred


def locate_sheet(workbook: openpyxl.Workbook):
    if "期末" not in workbook.sheetnames:
        raise ValueError("找不到工作表「期末」")
    return workbook["期末"]


def read_and_validate(path: Path, period_end: date) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = locate_sheet(workbook)
        headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=4, max_row=4))]
        if headers[:9] != EXPECTED_HEADERS:
            raise ValueError(
                "第 4 列欄位不符合預期；應為：" + "、".join(EXPECTED_HEADERS)
            )

        rows: list[dict[str, Any]] = []
        errors: list[dict[str, Any]] = []
        warnings: list[dict[str, Any]] = []
        keys: set[tuple[str, str]] = set()
        model_categories: dict[str, tuple[str, str]] = {}

        for row_number, values in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
            values = tuple(values[:9]) + (None,) * max(0, 9 - len(values))
            dealer_code = clean_text(values[0])

            # Empty rows and the three report summary rows are not source records.
            if not dealer_code or not dealer_code.startswith("TW"):
                continue

            row_errors: list[str] = []
            try:
                short_name = clean_text(values[1])
                category = clean_text(values[2])
                subcategory = clean_text(values[3])
                model_code = clean_text(values[4])
                salesperson = clean_text(values[8])

                if not re.fullmatch(r"TW\d{9}H", dealer_code):
                    row_errors.append("TW CODE 格式錯誤")
                for label, item in (
                    ("簡稱", short_name),
                    ("產品別", category),
                    ("產品別2", subcategory),
                    ("型號", model_code),
                    ("業務員", salesperson),
                ):
                    if not item:
                        row_errors.append(f"{label}不可空白")

                display_qty = parse_integer(values[5], "陳列", row_number, nullable=True)
                ending_qty = parse_integer(values[6], "期末", row_number, nullable=False)
                if display_qty is not None and display_qty < 0:
                    row_errors.append("陳列不可小於 0")

                excluded_text = clean_text(values[7])
                if excluded_text not in (None, "0", "不計"):
                    row_errors.append(f"不計欄位無法識別：{excluded_text}")
                is_excluded = excluded_text == "不計"

                if model_code:
                    key = (dealer_code, model_code)
                    if key in keys:
                        row_errors.append("同一月份的 TW CODE＋型號重複")
                    keys.add(key)
                    category_pair = (category or "", subcategory or "")
                    previous = model_categories.get(model_code)
                    if previous and previous != category_pair:
                        row_errors.append("相同型號出現不同產品分類")
                    model_categories[model_code] = category_pair

                if ending_qty is not None and ending_qty < 0:
                    warnings.append({
                        "row": row_number,
                        "message": f"期末數量為 {ending_qty}，請確認是否為盤點調整",
                    })

                if row_errors:
                    errors.append({"row": row_number, "message": "；".join(row_errors)})
                    continue

                rows.append({
                    "source_row": row_number,
                    "dealer_code": dealer_code,
                    "short_name": short_name,
                    "category": category,
                    "subcategory": subcategory,
                    "model_code": model_code,
                    "display_qty": display_qty,
                    "ending_qty": ending_qty,
                    "is_excluded": is_excluded,
                    "salesperson": salesperson,
                    "raw": [clean_text(value) for value in values[:9]],
                })
            except ValueError as exc:
                errors.append({"row": row_number, "message": str(exc)})

        if not rows and not errors:
            errors.append({"row": 0, "message": "工作表沒有可匯入的明細資料"})
        return rows, warnings, errors
    finally:
        workbook.close()


def cleanup_pending() -> None:
    now = datetime.now()
    expired: list[str] = []
    with pending_lock:
        for token, item in pending_imports.items():
            if (now - item.created_at).total_seconds() > TOKEN_TTL_SECONDS:
                expired.append(token)
        for token in expired:
            item = pending_imports.pop(token)
            item.path.unlink(missing_ok=True)


def check_existing_period(period_end: date) -> int:
    with db_connect() as connection:
        cursor = connection.cursor()
        return int(cursor.execute(
            "SELECT COUNT_BIG(*) FROM dbo.InventoryMonthEnd WHERE PeriodEnd = ?",
            period_end,
        ).fetchone()[0])


def import_to_database(item: PendingImport, replace_existing: bool) -> dict[str, Any]:
    connection = db_connect()
    connection.autocommit = False
    cursor = connection.cursor()
    try:
        existing = int(cursor.execute(
            "SELECT COUNT_BIG(*) FROM dbo.InventoryMonthEnd WITH (UPDLOCK, HOLDLOCK) WHERE PeriodEnd = ?",
            item.period_end,
        ).fetchone()[0])
        if existing and not replace_existing:
            raise ValueError(f"{item.period_end:%Y-%m} 已有 {existing:,} 筆資料，請勾選取代後再匯入")

        cursor.execute(
            """
            INSERT dbo.ImportBatch
                (PeriodEnd, SourceFileName, SourceSheetName, SourceRowCount,
                 ImportedRowCount, RejectedRowCount, ImportStatus)
            OUTPUT inserted.ImportBatchId
            VALUES (?, ?, N'期末', ?, 0, 0, 'Validated');
            """,
            item.period_end, item.original_name, len(item.rows),
        )
        batch_id = int(cursor.fetchone()[0])

        stage_values = [(
            batch_id, row["source_row"], *row["raw"], "Valid"
        ) for row in item.rows]
        cursor.fast_executemany = True
        cursor.executemany(
            """
            INSERT dbo.InventoryMonthEndStage
                (ImportBatchId, SourceRowNumber, DealerCodeRaw, ShortNameRaw,
                 CategoryRaw, SubcategoryRaw, ModelCodeRaw, DisplayQtyRaw,
                 EndingQtyRaw, ExcludedRaw, SalespersonRaw, ValidationStatus)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
            """,
            stage_values,
        )

        dealers = sorted({(r["dealer_code"], r["short_name"]) for r in item.rows})
        salespeople = sorted({r["salesperson"] for r in item.rows})
        categories = sorted({(r["category"], r["subcategory"]) for r in item.rows})
        products = sorted({(r["model_code"], r["category"], r["subcategory"]) for r in item.rows})

        cursor.fast_executemany = False
        for code, name in dealers:
            cursor.execute(
                """
                IF EXISTS (SELECT 1 FROM dbo.Dealer WHERE DealerCode = ?)
                    UPDATE dbo.Dealer SET ShortName = ?, IsActive = 1, ModifiedAt = sysdatetime()
                    WHERE DealerCode = ?;
                ELSE
                    INSERT dbo.Dealer (DealerCode, ShortName) VALUES (?, ?);
                """,
                code, name, code, code, name,
            )
        for name in salespeople:
            cursor.execute(
                "IF NOT EXISTS (SELECT 1 FROM dbo.Salesperson WHERE SalespersonName = ?) "
                "INSERT dbo.Salesperson (SalespersonName) VALUES (?);",
                name, name,
            )
        for category, subcategory in categories:
            cursor.execute(
                "IF NOT EXISTS (SELECT 1 FROM dbo.ProductCategory WHERE CategoryCode = ? AND SubcategoryName = ?) "
                "INSERT dbo.ProductCategory (CategoryCode, SubcategoryName) VALUES (?, ?);",
                category, subcategory, category, subcategory,
            )
        for model, category, subcategory in products:
            category_id = int(cursor.execute(
                "SELECT CategoryId FROM dbo.ProductCategory WHERE CategoryCode = ? AND SubcategoryName = ?",
                category, subcategory,
            ).fetchone()[0])
            existing_product = cursor.execute(
                "SELECT CategoryId FROM dbo.Product WHERE ModelCode = ?", model
            ).fetchone()
            if existing_product and int(existing_product[0]) != category_id:
                raise ValueError(f"型號 {model} 的既有產品分類與本次檔案不同")
            if not existing_product:
                cursor.execute(
                    "INSERT dbo.Product (ModelCode, CategoryId) VALUES (?, ?)", model, category_id
                )

        if existing:
            cursor.execute("DELETE dbo.InventoryMonthEnd WHERE PeriodEnd = ?", item.period_end)

        dealer_ids = {row.DealerCode: int(row.DealerId) for row in cursor.execute("SELECT DealerId, DealerCode FROM dbo.Dealer")}
        salesperson_ids = {row.SalespersonName: int(row.SalespersonId) for row in cursor.execute("SELECT SalespersonId, SalespersonName FROM dbo.Salesperson")}
        product_ids = {row.ModelCode: int(row.ProductId) for row in cursor.execute("SELECT ProductId, ModelCode FROM dbo.Product")}

        facts = [(
            item.period_end,
            dealer_ids[row["dealer_code"]],
            product_ids[row["model_code"]],
            salesperson_ids[row["salesperson"]],
            row["display_qty"],
            row["ending_qty"],
            int(row["is_excluded"]),
            batch_id,
            row["source_row"],
        ) for row in item.rows]
        cursor.fast_executemany = True
        cursor.executemany(
            """
            INSERT dbo.InventoryMonthEnd
                (PeriodEnd, DealerId, ProductId, SalespersonId, DisplayQuantity,
                 EndingQuantity, IsExcluded, ImportBatchId, SourceRowNumber)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);
            """,
            facts,
        )
        cursor.fast_executemany = False
        cursor.execute(
            "UPDATE dbo.InventoryMonthEndStage SET ValidationStatus = 'Loaded' WHERE ImportBatchId = ?",
            batch_id,
        )
        cursor.execute(
            """
            UPDATE dbo.ImportBatch
            SET ImportedRowCount = ?, RejectedRowCount = 0,
                ImportStatus = 'Completed', CompletedAt = sysdatetime()
            WHERE ImportBatchId = ?;
            """,
            len(item.rows), batch_id,
        )
        connection.commit()
        return {"batchId": batch_id, "importedRows": len(item.rows), "replacedRows": existing}
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


@app.get("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.get("/DBInput.html")
def db_input():
    return send_from_directory(BASE_DIR, "DBInput.html")


@app.get("/api/health")
def health():
    try:
        with db_connect() as connection:
            database_name = connection.cursor().execute("SELECT DB_NAME()").fetchone()[0]
        return jsonify({"ok": True, "database": database_name})
    except Exception as exc:
        return jsonify({"ok": False, "message": str(exc)}), 503


@app.get("/api/salespeople")
def get_salespeople():
    """Return active salespeople that have inventory records."""
    try:
        with db_connect() as connection:
            rows = connection.cursor().execute(
                """
                SELECT DISTINCT s.SalespersonId, s.SalespersonName
                FROM dbo.Salesperson AS s
                INNER JOIN dbo.InventoryMonthEnd AS i
                    ON i.SalespersonId = s.SalespersonId
                WHERE s.IsActive = 1
                ORDER BY s.SalespersonName;
                """
            ).fetchall()
        return jsonify({"ok": True, "items": [
            {"id": int(row.SalespersonId), "name": row.SalespersonName}
            for row in rows
        ]})
    except Exception as exc:
        return jsonify({"ok": False, "message": str(exc)}), 503


@app.get("/api/dealers")
def get_dealers():
    salesperson_id = request.args.get("salespersonId", type=int)
    if not salesperson_id:
        return jsonify({"ok": False, "message": "缺少 salespersonId"}), 400
    try:
        with db_connect() as connection:
            rows = connection.cursor().execute(
                """
                SELECT DISTINCT d.DealerId, d.DealerCode, d.ShortName
                FROM dbo.InventoryMonthEnd AS i
                INNER JOIN dbo.Dealer AS d ON d.DealerId = i.DealerId
                WHERE i.SalespersonId = ? AND d.IsActive = 1
                ORDER BY d.ShortName, d.DealerCode;
                """,
                salesperson_id,
            ).fetchall()
        return jsonify({"ok": True, "items": [
            {"id": int(row.DealerId), "code": row.DealerCode, "name": row.ShortName}
            for row in rows
        ]})
    except Exception as exc:
        return jsonify({"ok": False, "message": str(exc)}), 503


@app.get("/api/models")
def get_models():
    salesperson_id = request.args.get("salespersonId", type=int)
    dealer_id = request.args.get("dealerId", type=int)
    if not salesperson_id or not dealer_id:
        return jsonify({"ok": False, "message": "缺少 salespersonId 或 dealerId"}), 400
    try:
        with db_connect() as connection:
            rows = connection.cursor().execute(
                """
                WITH LatestPeriod AS (
                    SELECT MAX(PeriodEnd) AS PeriodEnd
                    FROM dbo.InventoryMonthEnd
                    WHERE DealerId = ?
                )
                SELECT p.ProductId, p.ModelCode, i.EndingQuantity
                FROM dbo.InventoryMonthEnd AS i
                INNER JOIN LatestPeriod AS lp ON lp.PeriodEnd = i.PeriodEnd
                INNER JOIN dbo.Product AS p ON p.ProductId = i.ProductId
                WHERE i.DealerId = ?
                  AND i.EndingQuantity > 0
                  AND p.IsActive = 1
                ORDER BY p.ModelCode;
                """,
                dealer_id, dealer_id,
            ).fetchall()
        return jsonify({"ok": True, "items": [
            {
                "id": int(row.ProductId),
                "modelCode": row.ModelCode,
                "currentQuantity": int(row.EndingQuantity),
            }
            for row in rows
        ]})
    except Exception as exc:
        return jsonify({"ok": False, "message": str(exc)}), 503


@app.post("/api/imports/validate")
def validate_import():
    cleanup_pending()
    uploaded = request.files.get("file")
    if uploaded is None or not uploaded.filename:
        return jsonify({"ok": False, "message": "請選擇 Excel 檔案"}), 400
    if Path(uploaded.filename).suffix.lower() != ".xlsx":
        return jsonify({"ok": False, "message": "只接受 .xlsx 檔案"}), 400

    safe_name = secure_filename(uploaded.filename) or "inventory.xlsx"
    temp = tempfile.NamedTemporaryFile(prefix="lgsaleout_", suffix="_" + safe_name, delete=False)
    temp_path = Path(temp.name)
    temp.close()
    try:
        uploaded.save(temp_path)
        period_end = parse_period(request.form.get("period"), uploaded.filename)
        rows, warnings, errors = read_and_validate(temp_path, period_end)
        if errors:
            temp_path.unlink(missing_ok=True)
            return jsonify({
                "ok": False,
                "message": "檔案檢查未通過",
                "errors": errors[:100],
                "errorCount": len(errors),
                "warnings": warnings[:100],
            }), 422

        token = secrets.token_urlsafe(32)
        pending = PendingImport(temp_path, uploaded.filename, period_end, rows, warnings)
        with pending_lock:
            pending_imports[token] = pending

        existing_rows = check_existing_period(period_end)
        return jsonify({
            "ok": True,
            "token": token,
            "period": period_end.isoformat(),
            "rowCount": len(rows),
            "dealerCount": len({r["dealer_code"] for r in rows}),
            "productCount": len({r["model_code"] for r in rows}),
            "displayTotal": sum(r["display_qty"] or 0 for r in rows),
            "endingTotal": sum(r["ending_qty"] for r in rows),
            "excludedTotal": sum(r["ending_qty"] for r in rows if r["is_excluded"]),
            "countedTotal": sum(r["ending_qty"] for r in rows if not r["is_excluded"]),
            "existingRows": existing_rows,
            "warnings": warnings[:100],
            "warningCount": len(warnings),
        })
    except (ValueError, openpyxl.utils.exceptions.InvalidFileException) as exc:
        temp_path.unlink(missing_ok=True)
        return jsonify({"ok": False, "message": str(exc)}), 400
    except Exception as exc:
        temp_path.unlink(missing_ok=True)
        return jsonify({"ok": False, "message": f"無法檢查檔案：{exc}"}), 500


@app.post("/api/imports/commit")
def commit_import():
    payload = request.get_json(silent=True) or {}
    token = payload.get("token")
    if not token:
        return jsonify({"ok": False, "message": "缺少匯入識別碼，請重新檢查檔案"}), 400
    with pending_lock:
        item = pending_imports.get(token)
    if item is None:
        return jsonify({"ok": False, "message": "檔案已逾時，請重新選擇並檢查"}), 410

    try:
        result = import_to_database(item, bool(payload.get("replaceExisting")))
        with pending_lock:
            pending_imports.pop(token, None)
        item.path.unlink(missing_ok=True)
        return jsonify({"ok": True, **result})
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 409
    except Exception as exc:
        return jsonify({"ok": False, "message": f"資料庫匯入失敗：{exc}"}), 500


@app.errorhandler(413)
def file_too_large(_error):
    return jsonify({"ok": False, "message": "檔案超過 20 MB 上限"}), 413


if __name__ == "__main__":
    lan_ip = local_network_ip()
    print("=" * 54)
    print(" LGSaleOut server is running")
    print(" Computer : http://127.0.0.1:8097")
    if lan_ip:
        print(f" Mobile   : http://{lan_ip}:8097")
        print(" Phone and computer must use the same Wi-Fi.")
    else:
        print(" Mobile   : LAN IP could not be detected")
    print("=" * 54, flush=True)
    app.run(host="0.0.0.0", port=8097, debug=False)
